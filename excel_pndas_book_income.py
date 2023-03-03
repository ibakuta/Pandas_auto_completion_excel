import pandas as pd
from openpyxl import load_workbook
import requests
from datetime import datetime
import openpyxl
import os


def open_excel_bank(file_bank):
    '''фильтруем данные из банковской выписки (по зачисленному доходу в валюте на счет)'''
    try:
        df = pd.read_excel(file_bank, header = 0, skiprows=19)
        df = df.loc[df['Сума_операції'] > 0, ['Дата i час операції', 'Сума_операції', 'Валюта_операції']]
        df['Дата i час операції'] = pd.to_datetime(df['Дата i час операції'], format = '%d.%m.%Y %H:%M:%S')
        df['Дата_операції'] = df['Дата i час операції'].dt.day
        df['Місяць_операції'] = df['Дата i час операції'].dt.month
        df['Рік_операції'] = df['Дата i час операції'].dt.year

        cols = df.columns.tolist()
        cols = cols[8:] + cols[:8]
        df = df[cols]
        df = df.sort_values(by=['Місяць_операції','Дата_операції'])
        return df

    except FileNotFoundError:
        print('File does not exist')

def save__excel_for_calc(df, file_book):
    '''сохраняем в книгу учета доходов на временную страницу для расчетов показателей'''
    with pd.ExcelWriter(file_book, engine='openpyxl', if_sheet_exists='new', mode='a') as writer:
        df.to_excel(writer, sheet_name="Sheet1", index=False)
        writer.save()

def calculation_excel_book(file_book):
    '''производим вычисления в книге учета доходов'''
    
    '''парсим дату по формату даты на сайте НБУ для загрузки курсов'''
    df = pd.read_excel(file_book, 'Sheet11', header=0, parse_dates=['Дата i час операції'])
    
    df['Дата i час операції'] = pd.to_datetime(df['Дата i час операції'], format = '%Y-%m-%d %H:%M:%S', errors='coerce') #Pass errors='coerce' to convert unparsable data to NaT (not a time)
    df['Дата i час операції'] = df['Дата i час операції'].dt.strftime('%Y%m%d')

    '''загружаем курсы'''
    rate = []
    for i in df['Дата i час операції'].tolist():
        url = f'https://bank.gov.ua/NBUStatService/v1/statdirectory/exchange?valcode=EUR&date=' + str(i) + '&json'
        resp = requests.get(url, headers = None)
        if resp.status_code == 200:
            data = resp.json()
            for i in data:
                rate.append(i['rate'])
        else:
            print('Сайт не доступен')
    df['Курс_НБУ'] = rate        

    '''расчитываем сумму дохода'''
    df['Сума_доходу'] = round(df['Сума_операції']* df['Курс_НБУ'],2)
    
    '''добавляем Subtotal and Grand Total'''
    container = []
    for label, i in df.groupby(['Місяць_операції', 'Рік_операції']):
        i.loc[f'{label[0]} {label[1]} Subtotal'] = i[['Сума_операції', 'Сума_доходу']].sum()
        container.append(i)
    
    df_summary = pd.concat(container)
    df_summary.loc['Grand Total'] = df[['Сума_операції', 'Сума_доходу']].sum()
    df_summary.fillna('')

    df_summary['Дата'] = pd.to_datetime(df['Дата_операції'].astype(int).astype(str)+'-'+df['Місяць_операції'].astype(int).astype(str)+'-'+df['Рік_операції'].astype(int).astype(str), format = '%d-%m-%Y').dt.strftime('%d-%m-%Y')
    
    df_summary = df_summary.drop('Дата_операції', axis=1)
    df_summary = df_summary.drop('Місяць_операції', axis=1)
    df_summary = df_summary.drop('Рік_операції', axis=1)
    df_summary = df_summary.drop('Дата i час операції', axis=1)

    cols = df_summary.columns.tolist()
    cols = cols[-1:] + cols[:-1]
    df_summary = df_summary[cols]

    print(df_summary)
    
    return df_summary


def save__excel_cumulative_total(df_summary, file_book):
    '''сохраняем в книгу учета доходов на основную страницу нарастающим итогом''' 
    book = load_workbook(file_book)
    writer = pd.ExcelWriter(file_book, engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    df_summary.to_excel(writer, sheet_name="Sheet1", startrow=writer.sheets['Sheet1'].max_row, index = False, header = False)
    writer.save()


def delete_Sheet11(file_book):
    '''удаляем промежуточную страницу Sheet11'''
    wb = openpyxl.load_workbook(file_book) #Open Excel-file
    sheet = wb.sheetnames #Got a list of all sheets in the file and drove it into a variable
    pfd = wb['Sheet11'] #We made the page we want to delete active, where [1] is the name of the page. It’s clear that you can make it a variable ;)
    wb.remove(pfd) #Deleting this page
    wb.save(file_book) #Saved file with changes (deleted page)

if __name__ == '__main__':
    df = open_excel_bank('statement_2021-10-01_2021-12-31_EUR.xlsx')
    save__excel_for_calc(df, 'книга доходів.xlsx')
    df_summary = calculation_excel_book('книга доходів.xlsx')
    save__excel_cumulative_total(df_summary,'книга доходів.xlsx')
    delete_Sheet11('книга доходів.xlsx')








